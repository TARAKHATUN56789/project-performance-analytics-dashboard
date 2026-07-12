"""
04_build_excel_dashboard.py
-----------------------------
Builds the Excel dashboard deliverable: a multi-sheet workbook with
raw data, KPI cards, summary tables (built with Excel formulas, not
hardcoded numbers), and native Excel charts.

Sheets:
  1. Executive Dashboard  - KPI cards + charts
  2. Project Data         - cleaned project table
  3. Task Data            - cleaned task table
  4. Dept Summary         - department-level formulas
  5. Manager Summary      - manager-level formulas
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

DATA_DIR = "../data"
OUT_PATH = "../excel/Project_Performance_Dashboard.xlsx"

# Google brand-inspired palette
BLUE = "4285F4"
RED = "EA4335"
YELLOW = "FBBC05"
GREEN = "34A853"
DARK = "202124"
LIGHT_GRAY = "F1F3F4"

projects = pd.read_csv(f"{DATA_DIR}/project_data_clean.csv")
tasks = pd.read_csv(f"{DATA_DIR}/task_data_clean.csv")

wb = Workbook()

# ---------------------------------------------------------------------------
# SHEET: Project Data
# ---------------------------------------------------------------------------
ws_proj = wb.active
ws_proj.title = "Project Data"
for r in dataframe_to_rows(projects, index=False, header=True):
    ws_proj.append(r)

header_fill = PatternFill("solid", fgColor=DARK)
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
for cell in ws_proj[1]:
    cell.fill = header_fill
    cell.font = header_font
ws_proj.freeze_panes = "A2"
for i, col in enumerate(projects.columns, 1):
    ws_proj.column_dimensions[get_column_letter(i)].width = max(14, len(col) + 2)

n_proj_rows = projects.shape[0] + 1  # +1 for header

# ---------------------------------------------------------------------------
# SHEET: Task Data
# ---------------------------------------------------------------------------
ws_task = wb.create_sheet("Task Data")
for r in dataframe_to_rows(tasks, index=False, header=True):
    ws_task.append(r)
for cell in ws_task[1]:
    cell.fill = header_fill
    cell.font = header_font
ws_task.freeze_panes = "A2"
for i, col in enumerate(tasks.columns, 1):
    ws_task.column_dimensions[get_column_letter(i)].width = max(14, len(col) + 2)

n_task_rows = tasks.shape[0] + 1

# Helper to find column letter by name
def col_letter(df, name):
    return get_column_letter(list(df.columns).index(name) + 1)

# ---------------------------------------------------------------------------
# SHEET: Dept Summary (uses formulas referencing Project Data sheet)
# ---------------------------------------------------------------------------
ws_dept = wb.create_sheet("Dept Summary")
depts = sorted(projects["Department"].unique())
ws_dept.append(["Department", "Total Projects", "Completed", "Delayed",
                "Success Rate %", "Total Budget", "Total Actual Cost",
                "Budget Utilization %"])
for cell in ws_dept[1]:
    cell.fill = header_fill
    cell.font = header_font

dept_col = col_letter(projects, "Department")
status_col = col_letter(projects, "Status")
budget_col = col_letter(projects, "Budget")
cost_col = col_letter(projects, "Actual_Cost")

for i, d in enumerate(depts, start=2):
    rng_dept = f"'Project Data'!${dept_col}$2:${dept_col}${n_proj_rows}"
    rng_status = f"'Project Data'!${status_col}$2:${status_col}${n_proj_rows}"
    rng_budget = f"'Project Data'!${budget_col}$2:${budget_col}${n_proj_rows}"
    rng_cost = f"'Project Data'!${cost_col}$2:${cost_col}${n_proj_rows}"
    ws_dept.append([
        d,
        f'=COUNTIF({rng_dept},A{i})',
        f'=COUNTIFS({rng_dept},A{i},{rng_status},"Completed")',
        f'=COUNTIFS({rng_dept},A{i},{rng_status},"Delayed")',
        f'=IFERROR(C{i}/(C{i}+D{i}),0)',
        f'=SUMIF({rng_dept},A{i},{rng_budget})',
        f'=SUMIF({rng_dept},A{i},{rng_cost})',
        f'=IFERROR(G{i}/F{i},0)'
    ])

for row in ws_dept.iter_rows(min_row=2, max_row=len(depts) + 1):
    row[4].number_format = "0.0%"
    row[7].number_format = "0.0%"
    row[5].number_format = "#,##0"
    row[6].number_format = "#,##0"

for i in range(1, 9):
    ws_dept.column_dimensions[get_column_letter(i)].width = 18

# ---------------------------------------------------------------------------
# SHEET: Manager Summary
# ---------------------------------------------------------------------------
ws_mgr = wb.create_sheet("Manager Summary")
mgrs = sorted(projects["Project_Manager"].unique())
ws_mgr.append(["Project Manager", "Total Projects", "Completed", "Delayed", "Success Rate %"])
for cell in ws_mgr[1]:
    cell.fill = header_fill
    cell.font = header_font

mgr_col = col_letter(projects, "Project_Manager")
for i, m in enumerate(mgrs, start=2):
    rng_mgr = f"'Project Data'!${mgr_col}$2:${mgr_col}${n_proj_rows}"
    rng_status = f"'Project Data'!${status_col}$2:${status_col}${n_proj_rows}"
    ws_mgr.append([
        m,
        f'=COUNTIF({rng_mgr},A{i})',
        f'=COUNTIFS({rng_mgr},A{i},{rng_status},"Completed")',
        f'=COUNTIFS({rng_mgr},A{i},{rng_status},"Delayed")',
        f'=IFERROR(C{i}/(C{i}+D{i}),0)'
    ])
for row in ws_mgr.iter_rows(min_row=2, max_row=len(mgrs) + 1):
    row[4].number_format = "0.0%"
for i in range(1, 6):
    ws_mgr.column_dimensions[get_column_letter(i)].width = 20

# ---------------------------------------------------------------------------
# SHEET: Executive Dashboard (KPI cards + charts)
# ---------------------------------------------------------------------------
ws_dash = wb.create_sheet("Executive Dashboard", 0)
ws_dash.sheet_view.showGridLines = False

ws_dash["B2"] = "PROJECT PERFORMANCE ANALYTICS — EXECUTIVE DASHBOARD"
ws_dash["B2"].font = Font(bold=True, size=16, color=DARK, name="Arial")
ws_dash.merge_cells("B2:J2")

kpi_defs = [
    ("Total Projects", f"=COUNTA('Project Data'!${dept_col}$2:${dept_col}${n_proj_rows})", BLUE),
    ("Completed", f'=COUNTIF(\'Project Data\'!${status_col}$2:${status_col}${n_proj_rows},"Completed")', GREEN),
    ("Delayed", f'=COUNTIF(\'Project Data\'!${status_col}$2:${status_col}${n_proj_rows},"Delayed")', RED),
    ("Success Rate %", "=D5/(D5+F5)", YELLOW),
]

positions = ["B4", "D4", "F4", "H4"]
for (label, formula, color), pos in zip(kpi_defs, positions):
    col = pos[0]
    row = int(pos[1:])
    label_cell = ws_dash[f"{col}{row}"]
    ws_dash.merge_cells(f"{col}{row}:{col}{row}")
    val_cell = ws_dash[f"{col}{row+1}"]
    label_cell.value = label
    label_cell.font = Font(size=10, color="666666", name="Arial")
    val_cell.value = formula
    val_cell.font = Font(size=22, bold=True, color=color, name="Arial")
    if "Rate" in label:
        val_cell.number_format = "0.0%"
    fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    for rr in range(row, row + 2):
        for cc in range(0, 1):
            pass

ws_dash.column_dimensions["A"].width = 3
for c in ["B", "D", "F", "H"]:
    ws_dash.column_dimensions[c].width = 20
for c in ["C", "E", "G", "I"]:
    ws_dash.column_dimensions[c].width = 3

# --- Charts on the dashboard, built from Dept Summary sheet ---
bar = BarChart()
bar.title = "Budget vs Actual Cost by Department"
bar.style = 10
bar.y_axis.title = "Amount"
data = Reference(ws_dept, min_col=6, max_col=7, min_row=1, max_row=len(depts) + 1)
cats = Reference(ws_dept, min_col=1, min_row=2, max_row=len(depts) + 1)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.width, bar.height = 16, 9
ws_dash.add_chart(bar, "B8")

pie = PieChart()
pie.title = "Projects by Status"
status_counts = projects["Status"].value_counts()
start_row = len(depts) + 5
ws_dept.cell(row=1, column=10, value="Status")
ws_dept.cell(row=1, column=11, value="Count")
for i, (k, v) in enumerate(status_counts.items(), start=2):
    ws_dept.cell(row=i, column=10, value=k)
    ws_dept.cell(row=i, column=11, value=v)
pie_data = Reference(ws_dept, min_col=11, max_col=11, min_row=1, max_row=len(status_counts) + 1)
pie_cats = Reference(ws_dept, min_col=10, min_row=2, max_row=len(status_counts) + 1)
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(pie_cats)
pie.width, pie.height = 12, 9
ws_dash.add_chart(pie, "B28")

line = LineChart()
line.title = "Success Rate % by Manager"
line.y_axis.title = "Success Rate"
mgr_data = Reference(ws_mgr, min_col=5, max_col=5, min_row=1, max_row=len(mgrs) + 1)
mgr_cats = Reference(ws_mgr, min_col=1, min_row=2, max_row=len(mgrs) + 1)
line.add_data(mgr_data, titles_from_data=True)
line.set_categories(mgr_cats)
line.width, line.height = 16, 9
ws_dash.add_chart(line, "J8")

wb.save(OUT_PATH)
print("Saved:", OUT_PATH)
